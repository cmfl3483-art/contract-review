"""
合规规则 Excel 批量导入服务
Compliance rules Excel batch import service
"""

import io
import json
import secrets
import uuid
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font
from fastapi import HTTPException

from app.core.redis_client import redis_client
from app.models.compliance import ComplianceRule, ComplianceRuleSet, RuleType, RuleSeverity
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, update

IMPORT_PREVIEW_TTL = 600                        # 10 分钟
IMPORT_RULES_LIMIT = 200
IMPORT_FILE_SIZE_LIMIT = 5 * 1024 * 1024        # 5 MB
IMPORT_EXCEL_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
IMPORT_REDIS_KEY_PREFIX = "compliance:import:preview:"

# 模板列定义（顺序固定）
TEMPLATE_COLUMNS = [
    ("rule_type",    "规则类型",   "枚举：number / name / description / file"),
    ("title",        "规则名称",   "1-100 字符"),
    ("requirement",  "规则正文",   "1-2000 字符"),
    ("severity",     "严重程度",   "枚举：must / should，缺省默认 must"),
    ("order",        "排序",       "整数，缺省默认 0"),
]


class ComplianceImportService:

    # ── 模板生成 ──────────────────────────────────────────────────────────────

    def generate_template(self) -> bytes:
        """
        生成标准 Excel 模板文件（bytes）。
        - 第 1 行：表头行（加粗）
        - 第 2 行：说明行（灰色背景 D9D9D9）
        - 第 3 行起：数据区域（含 rule_type / severity 下拉验证）
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "合规规则"

        # 表头行（加粗）
        for col_idx, (field, label, _) in enumerate(TEMPLATE_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = Font(bold=True)

        # 说明行（灰色背景 D9D9D9）
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for col_idx, (_, _, hint) in enumerate(TEMPLATE_COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=hint)
            cell.fill = gray_fill

        # 数据验证：rule_type 列（A 列，第 3 行起）
        dv_type = DataValidation(
            type="list",
            formula1='"number,name,description,file"',
            allow_blank=False,
            showDropDown=False,
        )
        ws.add_data_validation(dv_type)
        dv_type.sqref = "A3:A1048576"

        # 数据验证：severity 列（D 列，第 3 行起）
        dv_severity = DataValidation(
            type="list",
            formula1='"must,should"',
            allow_blank=True,
            showDropDown=False,
        )
        ws.add_data_validation(dv_severity)
        dv_severity.sqref = "D3:D1048576"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── 上传预览 ──────────────────────────────────────────────────────────────

    async def parse_and_preview(
        self,
        *,
        file_data: bytes,
        file_mime_type: str,
        file_size: int,
        rule_set_id: str,
        db: AsyncSession,
    ) -> dict:
        """
        解析 Excel 文件，全量校验，生成预览 token 存入 Redis。

        Returns:
            { "preview_session_token": str, "rules": [...], "total_count": int }

        Raises:
            HTTPException(404): rule_set_id 不存在
            HTTPException(409): 导入后总规则数超过 200 条
            HTTPException(422): 文件类型/大小/内容校验失败
        """
        # 1. 文件类型 / 大小校验
        if file_mime_type != IMPORT_EXCEL_MIME:
            raise HTTPException(422, detail={
                "code": "import_invalid_file",
                "message": "文件类型不合法，请上传 .xlsx 格式文件",
                "field": "file",
            })
        if file_size > IMPORT_FILE_SIZE_LIMIT:
            raise HTTPException(422, detail={
                "code": "import_invalid_file",
                "message": "文件大小超过 5 MB 限制",
                "field": "file",
            })

        # 2. rule_set 存在性校验
        rule_set = await self._get_rule_set(rule_set_id, db)

        # 3. 解析 Excel
        parsed_rows, errors = self._parse_excel(file_data)

        # 4. 空行校验
        if not parsed_rows and not errors:
            raise HTTPException(422, detail={
                "code": "import_empty_file",
                "message": "文件中未包含有效数据行（第三行起为空）",
            })

        # 5. 行级校验错误汇总
        if errors:
            raise HTTPException(422, detail={
                "code": "import_validation_failed",
                "message": f"共 {len(errors)} 行数据校验失败",
                "errors": errors,
            })

        # 6. 超 200 条校验
        if len(parsed_rows) > IMPORT_RULES_LIMIT:
            raise HTTPException(422, detail={
                "code": "import_too_many_rows",
                "message": f"单次导入最多 {IMPORT_RULES_LIMIT} 条规则，本次解析到 {len(parsed_rows)} 条",
            })

        # 7. 与现有规则数量合并校验
        current_count = await self._get_rule_count(rule_set_id, db)
        if current_count + len(parsed_rows) > IMPORT_RULES_LIMIT:
            raise HTTPException(409, detail={
                "code": "import_quota_exceeded",
                "message": (
                    f"导入后总规则数将超过 {IMPORT_RULES_LIMIT} 条上限。"
                    f"当前已有 {current_count} 条，本次导入 {len(parsed_rows)} 条。"
                ),
                "current_count": current_count,
                "import_count": len(parsed_rows),
                "limit": IMPORT_RULES_LIMIT,
            })

        # 8. 生成 token，存入 Redis
        token = secrets.token_urlsafe(32)
        redis_key = f"{IMPORT_REDIS_KEY_PREFIX}{token}"
        await redis_client.set(
            redis_key,
            json.dumps(parsed_rows, ensure_ascii=False),
            ex=IMPORT_PREVIEW_TTL,
        )

        return {
            "preview_session_token": token,
            "rules": parsed_rows,
            "total_count": len(parsed_rows),
        }

    # ── 确认导入 ──────────────────────────────────────────────────────────────

    async def confirm_import(
        self,
        *,
        rule_set_id: str,
        preview_session_token: str,
        db: AsyncSession,
    ) -> dict:
        """
        从 Redis 读取预览数据，在单个事务中批量写入 compliance_rules。

        Returns:
            { "imported_count": int, "rule_set_id": str }

        Raises:
            HTTPException(404): rule_set_id 不存在
            HTTPException(409): 竞态导致总规则数超过 200 条
            HTTPException(422): token 不存在或已过期
            HTTPException(500): 事务执行失败
        """
        # 1. 校验 token
        redis_key = f"{IMPORT_REDIS_KEY_PREFIX}{preview_session_token}"
        raw = await redis_client.get(redis_key)
        if raw is None:
            raise HTTPException(422, detail={
                "code": "import_preview_expired",
                "message": "预览会话已过期，请重新上传 Excel 文件",
            })

        # 2. 反序列化规则列表
        parsed_rows = json.loads(raw)

        # 3. rule_set 存在性校验
        rule_set = await self._get_rule_set(rule_set_id, db)

        # 4. 竞态保护：再次校验规则数量
        current_count = await self._get_rule_count(rule_set_id, db)
        if current_count + len(parsed_rows) > IMPORT_RULES_LIMIT:
            raise HTTPException(409, detail={
                "code": "import_quota_exceeded",
                "message": (
                    f"导入后总规则数将超过 {IMPORT_RULES_LIMIT} 条上限。"
                    f"当前已有 {current_count} 条，本次导入 {len(parsed_rows)} 条。"
                ),
                "current_count": current_count,
                "import_count": len(parsed_rows),
                "limit": IMPORT_RULES_LIMIT,
            })

        # 5. 单事务批量写入
        try:
            now = datetime.utcnow()
            for row in parsed_rows:
                rule = ComplianceRule(
                    rule_set_id=uuid.UUID(str(rule_set_id)),
                    rule_type=RuleType(row["rule_type"]),
                    title=row["title"],
                    requirement=row["requirement"],
                    severity=RuleSeverity(row.get("severity", "must")),
                    order=int(row.get("order", 0)),
                    created_at=now,
                    updated_at=now,
                )
                db.add(rule)

            # 更新 rule_set.updated_at
            await db.execute(
                update(ComplianceRuleSet)
                .where(ComplianceRuleSet.id == uuid.UUID(str(rule_set_id)))
                .values(updated_at=now)
            )
            await db.commit()
        except Exception:
            await db.rollback()
            raise HTTPException(500, detail={
                "code": "import_transaction_failed",
                "message": "导入失败，数据未写入，请稍后重试",
            })

        # 6. token 失效（一次性令牌，事务 commit 后立即删除）
        await redis_client.delete(redis_key)

        return {
            "imported_count": len(parsed_rows),
            "rule_set_id": str(rule_set_id),
        }

    # ── 私有辅助方法 ──────────────────────────────────────────────────────────

    def _parse_excel(self, file_data: bytes) -> tuple[list[dict], list[dict]]:
        """
        解析 Excel 文件，跳过第 1-2 行，从第 3 行起逐行读取。
        返回 (parsed_rows, errors)。
        errors 中每项格式：{ "row_number": int, "field": str, "message": str }
        """
        wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
        ws = wb.active

        parsed_rows = []
        errors = []

        for excel_row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            # 跳过全空行
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            row_errors = []
            rule_type_val = str(row[0]).strip() if row[0] is not None else ""
            title_val = str(row[1]).strip() if row[1] is not None else ""
            requirement_val = str(row[2]).strip() if row[2] is not None else ""
            severity_val = str(row[3]).strip() if row[3] is not None else ""
            order_val = row[4]

            # rule_type 校验
            if rule_type_val not in ("number", "name", "description", "file"):
                row_errors.append({
                    "row_number": excel_row_idx,
                    "field": "rule_type",
                    "message": (
                        f"第 {excel_row_idx} 行：rule_type 取值 '{rule_type_val}' 不合法，"
                        "必须为 number / name / description / file 之一"
                    ),
                })

            # title 校验
            if not (1 <= len(title_val) <= 100):
                row_errors.append({
                    "row_number": excel_row_idx,
                    "field": "title",
                    "message": f"第 {excel_row_idx} 行：title 长度必须为 1-100 字符",
                })

            # requirement 校验
            if not (1 <= len(requirement_val) <= 2000):
                row_errors.append({
                    "row_number": excel_row_idx,
                    "field": "requirement",
                    "message": f"第 {excel_row_idx} 行：requirement 长度必须为 1-2000 字符",
                })

            # severity 校验（缺省默认 must）
            if severity_val == "":
                severity_val = "must"
            elif severity_val not in ("must", "should"):
                row_errors.append({
                    "row_number": excel_row_idx,
                    "field": "severity",
                    "message": (
                        f"第 {excel_row_idx} 行：severity 取值 '{severity_val}' 不合法，"
                        "必须为 must / should 之一"
                    ),
                })

            # order 校验（缺省默认 0）
            if order_val is None or str(order_val).strip() == "":
                order_int = 0
            else:
                try:
                    order_int = int(order_val)
                except (ValueError, TypeError):
                    order_int = None
                    row_errors.append({
                        "row_number": excel_row_idx,
                        "field": "order",
                        "message": f"第 {excel_row_idx} 行：order 必须为整数",
                    })

            if row_errors:
                errors.extend(row_errors)
            else:
                parsed_rows.append({
                    "row_number": excel_row_idx,
                    "rule_type": rule_type_val,
                    "title": title_val,
                    "requirement": requirement_val,
                    "severity": severity_val,
                    "order": order_int,
                })

        return parsed_rows, errors

    async def _get_rule_set(self, rule_set_id: str, db: AsyncSession) -> ComplianceRuleSet:
        """查询 rule_set，不存在则抛 404。"""
        result = await db.execute(
            select(ComplianceRuleSet).where(
                ComplianceRuleSet.id == uuid.UUID(str(rule_set_id))
            )
        )
        rule_set = result.scalar_one_or_none()
        if rule_set is None:
            raise HTTPException(404, detail="规范集合不存在")
        return rule_set

    async def _get_rule_count(self, rule_set_id: str, db: AsyncSession) -> int:
        """返回指定 rule_set 当前的规则数量。"""
        result = await db.execute(
            select(func.count(ComplianceRule.id)).where(
                ComplianceRule.rule_set_id == uuid.UUID(str(rule_set_id))
            )
        )
        return result.scalar_one()
